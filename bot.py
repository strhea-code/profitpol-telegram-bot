import telebot
from telebot import types
from openpyxl import load_workbook
from datetime import datetime
from zoneinfo import ZoneInfo

from config import TOKEN, file_path

bot = telebot.TeleBot(TOKEN)

# пустой словарь с ответами пользователя
user_data = {}

# текст кнопок для выбора даты
TODAY_BUTTON = 'Сегодня'
MANUAL_DATE_BUTTON = 'Ввести другую дату'

# текст кнопок подтверждения
CONFIRM_BUTTON = 'Подтвердить'
CANCEL_BUTTON = 'Отмена'


# создание списка объектов
def load_object_list():
    workbook = load_workbook(file_path)
    sheet = workbook['Исходные данные']

    object_list = []

    for row in range(2, sheet.max_row + 1):
        value = sheet[f'M{row}'].value
        if value:
            object_list.append(value)

    workbook.close()
    return object_list


# создание списка наименований работ
def load_work_list():
    workbook = load_workbook(file_path)
    sheet = workbook['Исходные данные']

    work_list = []

    for row in range(2, sheet.max_row + 1):
        value = sheet[f'F{row}'].value
        if value:
            work_list.append(value)

    workbook.close()
    return work_list


# чтение разрешенных Telegram ID из листа "Доступ"
def load_allowed_users():
    workbook = load_workbook(file_path)
    sheet = workbook['Доступ']

    allowed_users = []

    for row in range(2, sheet.max_row + 1):
        telegram_id = sheet[f'B{row}'].value

        if telegram_id:
            allowed_users.append(int(telegram_id))

    workbook.close()
    return allowed_users


# функция проверки доступа
def is_allowed_user(user_id):
    allowed_users = load_allowed_users()
    return user_id in allowed_users


# функция получения ФИО по Telegram ID
def get_fio_by_user_id(user_id):
    workbook = load_workbook(file_path)
    sheet = workbook['Доступ']

    for row in range(2, sheet.max_row + 1):
        fio = sheet[f'A{row}'].value
        telegram_id = sheet[f'B{row}'].value

        if telegram_id and int(telegram_id) == user_id:
            workbook.close()
            return fio

    workbook.close()
    return None


# функция проверки даты
def validate_date(date_text):
    try:
        datetime.strptime(date_text, "%d.%m.%Y")
        return True
    except ValueError:
        return False


# функция проверки объема работ
def validate_volume(volume_text):
    try:
        volume_text = volume_text.replace(',', '.').strip()
        volume = float(volume_text)

        if volume > 0:
            return True
        return False
    except ValueError:
        return False


# функция приведения объема к числу
def normalize_volume(volume_text):
    volume_text = volume_text.replace(',', '.').strip()
    return float(volume_text)


# функция красивого вывода объема
def format_volume(volume):
    if volume.is_integer():
        return str(int(volume))
    return str(volume).replace('.', ',')


# функция формирования итогового текста
def build_summary(user_id):
    return (
        f"Проверьте введенные данные:\n\n"
        f"ФИО: {user_data[user_id]['fio']}\n"
        f"Объект: {user_data[user_id]['object']}\n"
        f"Дата: {user_data[user_id]['date']}\n"
        f"Наименование работ: {user_data[user_id]['work']}\n"
        f"Объем: {format_volume(user_data[user_id]['volume'])}"
    )


# функция поиска следующей пустой строки в листе "Начисления"
def find_next_row(sheet):
    for row in range(2, sheet.max_row + 2):
        if (
            sheet[f'A{row}'].value in (None, '') and
            sheet[f'B{row}'].value in (None, '') and
            sheet[f'D{row}'].value in (None, '') and
            sheet[f'I{row}'].value in (None, '') and
            sheet[f'K{row}'].value in (None, '')
        ):
            return row

    return sheet.max_row + 1


# функция записи данных в Excel
def save_to_excel(user_id):
    workbook = load_workbook(file_path)
    sheet = workbook['Начисления']

    next_row = find_next_row(sheet)

    sheet[f'A{next_row}'] = user_data[user_id]['fio']
    sheet[f'B{next_row}'] = user_data[user_id]['object']
    sheet[f'D{next_row}'] = user_data[user_id]['date']
    sheet[f'I{next_row}'] = user_data[user_id]['work']
    sheet[f'K{next_row}'] = user_data[user_id]['volume']

    workbook.save(file_path)
    workbook.close()


# команда для просмотра своего Telegram ID
@bot.message_handler(commands=['myid'])
def myid_command(message):
    bot.send_message(message.chat.id, f"Ваш Telegram ID: {message.from_user.id}")


# обработчик команды /start
@bot.message_handler(commands=['start'])
def start_command(message):
    user_id = message.from_user.id

    if not is_allowed_user(user_id):
        bot.send_message(
            message.chat.id,
            "У вас нет доступа к этому боту.\n"
            "Отправьте команду /myid и передайте ваш Telegram ID заказчику."
        )
        return

    fio = get_fio_by_user_id(user_id)

    if not fio:
        bot.send_message(message.chat.id, "Не удалось определить ваше ФИО по Telegram ID.")
        return

    # сохраняем ФИО автоматически
    user_data[user_id] = {}
    user_data[user_id]["fio"] = fio

    bot.send_message(message.chat.id, f"Здравствуйте, {fio}")

    object_list = load_object_list()

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

    # создание кнопок с объектами
    for obj in object_list:
        button = types.KeyboardButton(obj)
        markup.add(button)

    bot.send_message(message.chat.id, "Выберите объект:", reply_markup=markup)


# реакция на сообщения пользователя
@bot.message_handler(func=lambda message: True)
def handle_message(message):
    user_id = message.from_user.id
    text = message.text.strip()

    if not is_allowed_user(user_id):
        bot.send_message(
            message.chat.id,
            "У вас нет доступа к этому боту.\n"
            "Отправьте команду /myid и передайте ваш Telegram ID заказчику."
        )
        return

    object_list = load_object_list()
    work_list = load_work_list()

    # если пользователь выбрал объект
    if text in object_list:

        if user_id not in user_data or "fio" not in user_data[user_id]:
            bot.send_message(message.chat.id, "Сначала нажмите /start")
            return

        user_data[user_id]["object"] = text

        bot.send_message(message.chat.id, f"Вы выбрали объект: {text}")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(
            types.KeyboardButton(TODAY_BUTTON),
            types.KeyboardButton(MANUAL_DATE_BUTTON)
        )

        bot.send_message(message.chat.id, "Выберите дату:", reply_markup=markup)
        return

    # если пользователь нажал "Сегодня"
    if text == TODAY_BUTTON:

        if user_id not in user_data or "object" not in user_data[user_id]:
            bot.send_message(message.chat.id, "Сначала выберите объект.")
            return

        today = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")
        user_data[user_id]["date"] = today

        bot.send_message(message.chat.id, f"Дата записана: {today}")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

        for work in work_list:
            button = types.KeyboardButton(work)
            markup.add(button)

        bot.send_message(message.chat.id, "Выберите наименование работ:", reply_markup=markup)
        return

    # если пользователь хочет ввести дату вручную
    if text == MANUAL_DATE_BUTTON:

        if user_id not in user_data or "object" not in user_data[user_id]:
            bot.send_message(message.chat.id, "Сначала выберите объект.")
            return

        bot.send_message(message.chat.id, "Введите дату в формате ДД.ММ.ГГГГ")
        return

    # если пользователь вводит дату вручную
    if user_id in user_data and "object" in user_data[user_id] and "date" not in user_data[user_id]:

        if validate_date(text):
            user_data[user_id]["date"] = text

            bot.send_message(message.chat.id, f"Дата записана: {text}")

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

            for work in work_list:
                button = types.KeyboardButton(work)
                markup.add(button)

            bot.send_message(message.chat.id, "Выберите наименование работ:", reply_markup=markup)
            return

        bot.send_message(
            message.chat.id,
            "❌ Дата введена некорректно.\nВведите дату в формате ДД.ММ.ГГГГ\nНапример: 12.03.2026"
        )
        return

    # если пользователь выбрал наименование работ
    if text in work_list:

        if user_id not in user_data or "date" not in user_data[user_id]:
            bot.send_message(message.chat.id, "Сначала выберите дату.")
            return

        user_data[user_id]["work"] = text

        bot.send_message(message.chat.id, f"Вы выбрали: {text}")
        bot.send_message(message.chat.id, "Введите объем выполненных работ:")
        return

    # если пользователь вводит объем работ
    if user_id in user_data and "work" in user_data[user_id] and "volume" not in user_data[user_id]:

        if validate_volume(text):

            user_data[user_id]["volume"] = normalize_volume(text)

            summary = build_summary(user_id)

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(
                types.KeyboardButton(CONFIRM_BUTTON),
                types.KeyboardButton(CANCEL_BUTTON)
            )

            bot.send_message(message.chat.id, summary, reply_markup=markup)
            bot.send_message(message.chat.id, "Подтвердите отправку данных или отмените.")
            return

        bot.send_message(
            message.chat.id,
            "❌ Объем введен некорректно.\nВведите только число.\nМожно использовать точку или запятую.\nНапример: 12 или 12.5 или 12,5"
        )
        return

    # если пользователь нажал "Подтвердить"
    if text == CONFIRM_BUTTON:

        if user_id not in user_data or "volume" not in user_data[user_id]:
            bot.send_message(message.chat.id, "Нет данных для подтверждения. Нажмите /start")
            return

        try:
            save_to_excel(user_id)
            bot.send_message(message.chat.id, "✅ Данные успешно записаны в Excel.")
            del user_data[user_id]

        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при записи в Excel: {e}")

        return

    # если пользователь нажал "Отмена"
    if text == CANCEL_BUTTON:

        if user_id in user_data:
            del user_data[user_id]

        bot.send_message(message.chat.id, "❌ Ввод данных отменен. Нажмите /start, чтобы начать заново.")
        return

    bot.send_message(message.chat.id, "Пожалуйста, выберите вариант кнопкой или нажмите /start")


bot.infinity_polling(skip_pending=True)